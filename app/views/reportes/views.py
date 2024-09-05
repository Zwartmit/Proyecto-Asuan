import openpyxl
from openpyxl.drawing.image import Image
from django.shortcuts import render
from django.http import HttpResponse
from app.forms import ReporteForm
from datetime import datetime
from app.models import *
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.utils.decorators import method_decorator
from django.views.decorators.cache import never_cache
from django.contrib.auth.decorators import login_required

def get(self, request, *args, **kwargs):
    contexto = {
        'titulo': 'Gestión de bases de datos'
    }
    return render(request, 'backup.html', contexto)

@login_required
@never_cache
def reporte_selector(request):
    if request.method == 'POST':
        tipo_reporte = request.POST.get('tipo_reporte')
        formato = request.POST.get('formato')
        
        if formato == 'excel':
            if tipo_reporte == 'categoria':
                return export_categorias_excel(request)
            elif tipo_reporte == 'marca':
                return export_marcas_excel(request)
            elif tipo_reporte == 'presentacion':
                return export_presentaciones_excel(request)
            elif tipo_reporte == 'producto':
                return export_productos_excel(request)
            elif tipo_reporte == 'plato':
                return export_platos_excel(request)
            elif tipo_reporte == 'mesero':
                return export_meseros_excel(request)
            elif tipo_reporte == 'cliente':
                return export_clientes_excel(request)
            elif tipo_reporte == 'administrador':
                return export_administradores_excel(request)
            elif tipo_reporte == 'operador':
                return export_operadores_excel(request)
        elif formato == 'pdf':
            if tipo_reporte == 'categoria':
                return export_categorias_pdf(request)
            elif tipo_reporte == 'marca':
                return export_marcas_pdf(request)
            elif tipo_reporte == 'presentacion':
                return export_presentaciones_pdf(request)
            elif tipo_reporte == 'producto':
                return export_productos_pdf(request)
            elif tipo_reporte == 'plato':
                return export_platos_pdf(request)
            elif tipo_reporte == 'mesero':
                return export_meseros_pdf(request)
            elif tipo_reporte == 'cliente':
                return export_clientes_pdf(request)
            elif tipo_reporte == 'administrador':
                return export_administradores_pdf(request)
            elif tipo_reporte == 'operador':
                return export_operadores_pdf(request)
    else:
        form = ReporteForm() 

    contexto = {
        'titulo': 'Generar reportes',
    }
    
    return render(request, 'reportes.html', contexto)
    return render(request, 'reportes.html', {'form': form})

################################################## Categorias ##################################################
def export_categorias_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de categorías"
    
    # Estilo
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    green_fill = PatternFill(start_color="04644B'", end_color="04644B", fill_type="solid")
    white_font = Font(color="FFFFFF")
    
    # Ajustar ancho de columnas a todas las columnas
    column_width = 20  # Ancho deseado
    for col in range(1, 11):  # Puedes ajustar el número de columnas según sea necesario
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = column_width

    # Ajustar alto de las filas
    ws.row_dimensions[1].height = 60  # Alto para la fila del logo
    ws.row_dimensions[5].height = 20  # Alto para el título
    ws.row_dimensions[7].height = 20  # Alto para los encabezados

    # Logo en la parte superior izquierda (ajustar la ruta)
    img = Image('app/views/reportes/logo_asuan.png')  # Reemplaza con la ruta correcta a tu logo
    img.width = 150  # Ajustar el ancho del logo
    img.height = 70  # Ajustar el alto del logo
    ws.add_image(img, 'B1')  # Posiciona la imagen en la celda A1

    # Título en la parte superior central (Centrado y combinando celdas)
    ws.merge_cells('A5:C5')
    ws['A5'] = "Reporte de Categorías"
    ws['A5'].font = Font(size=14, bold=True)
    ws['A5'].alignment = center_alignment

    # Fecha en la parte superior derecha
    ws.merge_cells('E3:F3')
    fecha = datetime.now().strftime("%d/%m/%Y")
    ws['E3'] = f"Fecha:\n{fecha}"
    ws['E3'].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    ws['E3'].font = bold_font

    # Encabezados de la tabla
    headers = ['ID', 'Categoría', 'Estado']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_num)
        cell.value = header
        cell.fill = green_fill
        cell.font = white_font
        cell.alignment = center_alignment
        cell.border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))

    # Agregar los datos de las categorías
    categorias = Categoria.objects.all()
    for row_num, categoria in enumerate(categorias, 8):  # Comienza en la fila 8
        ws.cell(row=row_num, column=1, value=categoria.id)
        ws.cell(row=row_num, column=2, value=categoria.categoria)
        ws.cell(row=row_num, column=3, value='Activo' if categoria.estado else 'Inactivo')
        
        for col_num in range(1, 4):  # Aplica el borde a las celdas de datos
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = center_alignment
            cell.border = Border(left=Side(style='thin'), 
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))

    # Generar archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=categorias.xlsx'
    wb.save(response)
    return response

def export_categorias_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Fecha en la parte superior izquierda
    fecha = datetime.now().strftime("%Y-%m-%d %H:%M")
    p.drawString(30, height - 50, f"Fecha: {fecha}")

    # Título en la parte superior central
    p.setFont("Helvetica-Bold", 16)
    p.drawString((width / 2) - 80, height - 50, "Reporte de Categorías")

    # Encabezados
    p.setFont("Helvetica-Bold", 12)
    p.drawString(100, height - 150, 'ID')
    p.drawString(200, height - 150, 'Categoría')
    p.drawString(300, height - 150, 'Estado')

    # Agregando los datos
    y = height - 170
    categorias = Categoria.objects.all()
    p.setFont("Helvetica", 10)
    for categoria in categorias:
        p.drawString(100, y, str(categoria.id))
        p.drawString(200, y, categoria.categoria)
        p.drawString(300, y, 'Activo' if categoria.estado else 'Inactivo')
        y -= 20

    # Guardar página
    p.showPage()
    p.save()

    # Retornar el PDF
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=categorias.pdf'
    return response

################################################## Marcas ##################################################
def export_marcas_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Marcas"

    headers = ['ID', 'Marca', 'Estado']
    ws.append(headers)

    marcas = Marca.objects.all()
    for marca in marcas:
        ws.append([
            marca.id,
            marca.marca,
            'Activo' if marca.estado else 'Inactivo',
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=marcas.xlsx'
    wb.save(response)
    return response

def export_marcas_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    p.drawString(100, height - 100, 'ID')
    p.drawString(200, height - 100, 'Marca')
    p.drawString(300, height - 100, 'Estado')

    y = height - 120
    marcas = Marca.objects.all()
    for marca in marcas:
        p.drawString(100, y, str(marca.id))
        p.drawString(200, y, marca.marca)
        p.drawString(300, y, 'Activo' if marca.estado else 'Inactivo')
        y -= 20

    p.showPage()
    p.save()

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=marcas.pdf'
    return response

################################################## Presentaciones ##################################################
def export_presentaciones_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Presentaciones"

    headers = ['ID', 'Presentación', 'Unidad de medida', 'Estado']
    ws.append(headers)

    presentaciones = Presentacion.objects.all()
    for presentacion in presentaciones:
        ws.append([
            presentacion.id,
            presentacion.presentacion,
            presentacion.get_unidad_medida_display(),
            'Activo' if presentacion.estado else 'Inactivo',
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=presentaciones.xlsx'
    wb.save(response)
    return response

def export_presentaciones_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    p.drawString(100, height - 100, 'ID')
    p.drawString(200, height - 100, 'Presentación')
    p.drawString(300, height - 100, 'Unidad de medida')
    p.drawString(400, height - 100, 'Estado')

    y = height - 120
    presentaciones = Presentacion.objects.all()
    for presentacion in presentaciones:
        p.drawString(100, y, str(presentacion.id))
        p.drawString(200, y, presentacion.presentacion)
        p.drawString(300, y, presentacion.get_unidad_medida_display())
        p.drawString(400, y, 'Activo' if presentacion.estado else 'Inactivo')
        y -= 20

    p.showPage()
    p.save()

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=presentaciones.pdf'
    return response

################################################## Productos ##################################################
def export_productos_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    headers = ['ID', 'Producto', 'Cantidad', 'Valor', 'Estado', 'Categoría', 'Marca', 'Presentación']
    ws.append(headers)

    productos = Producto.objects.all()
    for producto in productos:
        ws.append([
            producto.id,
            producto.producto,
            producto.cantidad,
            producto.valor,
            'Activo' if producto.estado else 'Inactivo',
            producto.id_categoria.categoria,
            producto.id_marca.marca,
            producto.id_presentacion.presentacion,
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=productos.xlsx'
    wb.save(response)
    return response

def export_productos_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, height - 50, 'ID')
    p.drawString(150, height - 50, 'Producto')
    p.drawString(250, height - 50, 'Cantidad')
    p.drawString(350, height - 50, 'Valor')
    p.drawString(450, height - 50, 'Estado')
    p.drawString(550, height - 50, 'Categoría')
    p.drawString(650, height - 50, 'Marca')
    p.drawString(750, height - 50, 'Presentación')

    y = height - 70
    productos = Producto.objects.all()
    for producto in productos:
        p.setFont("Helvetica", 10)
        p.drawString(50, y, str(producto.id))
        p.drawString(150, y, producto.producto)
        p.drawString(250, y, str(producto.cantidad))
        p.drawString(350, y, str(producto.valor))
        p.drawString(450, y, 'Activo' if producto.estado else 'Inactivo')
        p.drawString(550, y, producto.id_categoria.categoria)
        p.drawString(650, y, producto.id_marca.marca)
        p.drawString(750, y, producto.id_presentacion.presentacion)
        y -= 20

    p.showPage()
    p.save()
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=productos.pdf'
    return response

################################################## Platos ##################################################
def export_platos_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Platos"

    headers = ['ID', 'Nombre del plato', 'Descripción', 'Valor', 'Estado']
    ws.append(headers)

    platos = Plato.objects.all()
    for plato in platos:
        ws.append([
            plato.id,
            plato.plato,
            plato.descripcion,
            plato.valor,
            'Activo' if plato.estado else 'Inactivo',
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=platos.xlsx'
    wb.save(response)
    return response

def export_platos_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    p.drawString(100, height - 100, 'ID')
    p.drawString(200, height - 100, 'Nombre del plato')
    p.drawString(300, height - 100, 'Descripción')
    p.drawString(400, height - 100, 'Valor')
    p.drawString(500, height - 100, 'Estado')

    y = height - 120
    platos = Plato.objects.all()
    for plato in platos:
        p.drawString(100, y, str(plato.id))
        p.drawString(200, y, plato.plato)
        p.drawString(300, y, plato.descripcion)
        p.drawString(400, y, str(plato.valor))
        p.drawString(500, y, 'Activo' if plato.estado else 'Inactivo')
        y -= 20

    p.showPage()
    p.save()

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=platos.pdf'
    return response

################################################## Meseros ##################################################
def export_meseros_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Meseros"

    headers = ['ID', 'Nombre', 'Tipo de documento', 'Número de documento', 'Email', 'Prefijo telefónico', 'Teléfono']
    ws.append(headers)

    meseros = Mesero.objects.all()
    for mesero in meseros:
        ws.append([
            mesero.id,
            mesero.nombre,
            mesero.get_tipo_documento_display(),
            mesero.numero_documento,
            mesero.email,
            mesero.pais_telefono,
            mesero.telefono,
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=meseros.xlsx'
    wb.save(response)
    return response

def export_meseros_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    p.drawString(100, height - 100, 'ID')
    p.drawString(200, height - 100, 'Nombre')
    p.drawString(300, height - 100, 'Tipo de documento')
    p.drawString(400, height - 100, 'Número de documento')
    p.drawString(500, height - 100, 'Email')
    p.drawString(600, height - 100, 'Prefijo telefónico')
    p.drawString(700, height - 100, 'Teléfono')

    y = height - 120
    meseros = Mesero.objects.all()
    for mesero in meseros:
        p.drawString(100, y, str(mesero.id))
        p.drawString(200, y, mesero.nombre)
        p.drawString(300, y, mesero.get_tipo_documento_display())
        p.drawString(400, y, str(mesero.numero_documento))
        p.drawString(500, y, mesero.email)
        p.drawString(600, y, mesero.pais_telefono)
        p.drawString(700, y, str(mesero.telefono))
        y -= 20

    p.showPage()
    p.save()

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=meseros.pdf'
    return response

################################################## Clientes ##################################################
def export_clientes_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"

    headers = ['ID', 'Nombre', 'Tipo de documento', 'Número de documento', 'Email', 'Prefijo telefónico', 'Teléfono']
    ws.append(headers)

    clientes = Cliente.objects.all()
    for cliente in clientes:
        ws.append([
            cliente.id,
            cliente.nombre,
            cliente.get_tipo_documento_display(),
            cliente.numero_documento,
            cliente.email,
            cliente.pais_telefono,
            cliente.telefono,
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=clientes.xlsx'
    wb.save(response)
    return response

def export_clientes_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    p.drawString(100, height - 100, 'ID')
    p.drawString(200, height - 100, 'Nombre')
    p.drawString(300, height - 100, 'Tipo de documento')
    p.drawString(400, height - 100, 'Número de documento')
    p.drawString(500, height - 100, 'Email')
    p.drawString(600, height - 100, 'Prefijo telefónico')
    p.drawString(700, height - 100, 'Teléfono')

    y = height - 120
    clientes = Cliente.objects.all()
    for cliente in clientes:
        p.drawString(100, y, str(cliente.id))
        p.drawString(200, y, cliente.nombre)
        p.drawString(300, y, cliente.get_tipo_documento_display())
        p.drawString(400, y, str(cliente.numero_documento))
        p.drawString(500, y, cliente.email)
        p.drawString(600, y, cliente.pais_telefono)
        p.drawString(700, y, str(cliente.telefono))
        y -= 20

    p.showPage()
    p.save()

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=clientes.pdf'
    return response

################################################## Administradores ##################################################
def export_administradores_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Administradores"

    headers = ['ID', 'Nombre', 'Tipo de documento', 'Número de documento', 'Teléfono']
    ws.append(headers)

    administradores = Administrador.objects.all()
    for administrador in administradores:
        ws.append([
            administrador.id,
            administrador.nombre,
            administrador.get_tipo_documento_display(),
            administrador.numero_documento,
            administrador.telefono,
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=administradores.xlsx'
    wb.save(response)
    return response

def export_administradores_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    p.drawString(100, height - 100, 'ID')
    p.drawString(200, height - 100, 'Nombre')
    p.drawString(300, height - 100, 'Tipo de documento')
    p.drawString(400, height - 100, 'Número de documento')
    p.drawString(500, height - 100, 'Teléfono')

    y = height - 120
    administradores = Administrador.objects.all()
    for administrador in administradores:
        p.drawString(100, y, str(administrador.id))
        p.drawString(200, y, administrador.nombre)
        p.drawString(300, y, administrador.get_tipo_documento_display())
        p.drawString(400, y, str(administrador.numero_documento))
        p.drawString(500, y, str(administrador.telefono))
        y -= 20

    p.showPage()
    p.save()

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=administradores.pdf'
    return response

################################################## Operadores ##################################################
def export_operadores_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Operadores"

    headers = ['ID', 'Nombre', 'Tipo de documento', 'Número de documento', 'Teléfono']
    ws.append(headers)

    operadores = Operador.objects.all()
    for operador in operadores:
        ws.append([
            operador.id,
            operador.nombre,
            operador.get_tipo_documento_display(),
            operador.numero_documento,
            operador.telefono,
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=operadores.xlsx'
    wb.save(response)
    return response
def export_operadores_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    p.drawString(100, height - 100, 'ID')
    p.drawString(200, height - 100, 'Nombre')
    p.drawString(300, height - 100, 'Tipo de documento')
    p.drawString(400, height - 100, 'Número de documento')
    p.drawString(500, height - 100, 'Teléfono')

    y = height - 120
    operadores = Operador.objects.all()
    for operador in operadores:
        p.drawString(100, y, str(operador.id))
        p.drawString(200, y, operador.nombre)
        p.drawString(300, y, operador.get_tipo_documento_display())
        p.drawString(400, y, str(operador.numero_documento))
        p.drawString(500, y, str(operador.telefono))
        y -= 20

    p.showPage()
    p.save()

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=operadores.pdf'
    return response
