from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from django.views.decorators.cache import never_cache
from django.contrib.auth.decorators import login_required
from datetime import datetime
from app.models import *

################################################## Categorias ##################################################
@login_required
@never_cache
def export_categorias_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de categorías"
    
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    green_fill = PatternFill(start_color="04644B", end_color="04644B", fill_type="solid")
    white_font = Font(color="FFFFFF")
    medium_border = Border(left=Side(style='medium'), 
                         right=Side(style='medium'), 
                         top=Side(style='medium'), 
                         bottom=Side(style='medium'))

    column_width = 20  
    for col in range(2, 5): 
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = column_width

    ws.row_dimensions[2].height = 38 
    ws.row_dimensions[3].height = 23  
    ws.row_dimensions[5].height = 20

    img = Image('app/views/reportes/logo_asuan.png') 
    img.width = 140  
    img.height = 50  
    ws.add_image(img, 'C2')
    ws.merge_cells('B2:D2')
    ws['B2'].alignment = center_alignment
    ws['B2'].border = medium_border

    ws.merge_cells('B2:D2')
    ws['B2'].alignment = center_alignment
    ws.merge_cells('B3:D3')
    ws['B3'] = "Reporte de Categorías"
    ws['B3'].font = Font(size=14, bold=True)
    ws['B3'].alignment = center_alignment
    ws['B3'].border = medium_border
    ws['C3'].border = medium_border
    ws['D3'].border = medium_border

    ws.merge_cells('B4:D4')
    fecha = datetime.now().strftime("%d/%m/%Y")
    ws['B4'] = f"Fecha: {fecha}"
    ws['B4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws['B4'].border = medium_border
    ws['C4'].border = medium_border
    ws['D4'].border = medium_border

    headers = ['ID', 'Categoría', 'Estado']
    for col_num, header in enumerate(headers, 2):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.fill = green_fill
        cell.font = white_font
        cell.alignment = center_alignment
        cell.border = medium_border

    categorias = Categoria.objects.all()
    for row_num, categoria in enumerate(categorias, 6):  
        ws.cell(row=row_num, column=2, value=categoria.id)
        ws.cell(row=row_num, column=3, value=categoria.categoria)
        ws.cell(row=row_num, column=4, value='Activo' if categoria.estado else 'Inactivo')
        
        for col_num in range(2, 5):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = center_alignment
            cell.border = medium_border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte de categorías.xlsx'
    wb.save(response)
    return response

################################################## Marcas ##################################################
@login_required
@never_cache
def export_marcas_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de marcas"

    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    green_fill = PatternFill(start_color="04644B", end_color="04644B", fill_type="solid")
    white_font = Font(color="FFFFFF")
    medium_border = Border(left=Side(style='medium'), 
                         right=Side(style='medium'), 
                         top=Side(style='medium'), 
                         bottom=Side(style='medium'))

    column_width = 20  
    for col in range(2, 5): 
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = column_width

    ws.row_dimensions[2].height = 38 
    ws.row_dimensions[3].height = 23  
    ws.row_dimensions[5].height = 20

    img = Image('app/views/reportes/logo_asuan.png') 
    img.width = 140  
    img.height = 50  
    ws.add_image(img, 'C2')
    ws.merge_cells('B2:D2')
    ws['B2'].alignment = center_alignment
    ws['B2'].border = medium_border

    ws.merge_cells('B2:D2')
    ws['B2'].alignment = center_alignment
    ws.merge_cells('B3:D3')
    ws['B3'] = "Reporte de Marcas"
    ws['B3'].font = Font(size=14, bold=True)
    ws['B3'].alignment = center_alignment
    ws['B3'].border = medium_border
    for col in range(3, 5):
        ws.cell(row=3, column=col).border = medium_border

    ws.merge_cells('B4:D4')
    fecha = datetime.now().strftime("%d/%m/%Y")
    ws['B4'] = f"Fecha: {fecha}"
    ws['B4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws['B4'].border = medium_border
    for col in range(3, 5):
        ws.cell(row=4, column=col).border = medium_border

    headers = ['ID', 'Marca', 'Estado']
    for col_num, header in enumerate(headers, 2):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.fill = green_fill
        cell.font = white_font
        cell.alignment = center_alignment
        cell.border = medium_border

    marcas = Marca.objects.all()
    for row_num, marca in enumerate(marcas, 6):  
        ws.cell(row=row_num, column=2, value=marca.id)
        ws.cell(row=row_num, column=3, value=marca.marca)
        ws.cell(row=row_num, column=4, value='Activo' if marca.estado else 'Inactivo')
        
        for col_num in range(2, 5):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = center_alignment
            cell.border = medium_border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte de marcas.xlsx'
    wb.save(response)
    return response

################################################## Presentaciones ##################################################
@login_required
@never_cache
def export_presentaciones_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de presentaciones"

    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    green_fill = PatternFill(start_color="04644B", end_color="04644B", fill_type="solid")
    white_font = Font(color="FFFFFF")
    medium_border = Border(left=Side(style='medium'), 
                         right=Side(style='medium'), 
                         top=Side(style='medium'), 
                         bottom=Side(style='medium'))

    column_width = 20  
    for col in range(2, 5): 
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = column_width

    ws.row_dimensions[2].height = 38 
    ws.row_dimensions[3].height = 23  
    ws.row_dimensions[5].height = 20

    img = Image('app/views/reportes/logo_asuan.png') 
    img.width = 140  
    img.height = 50  
    ws.add_image(img, 'C2')
    ws.merge_cells('B2:D2')
    ws['B2'].alignment = center_alignment
    ws['B2'].border = medium_border

    ws.merge_cells('B2:D2')
    ws['B2'].alignment = center_alignment
    ws.merge_cells('B3:D3')
    ws['B3'] = "Reporte de Presentaciones"
    ws['B3'].font = Font(size=14, bold=True)
    ws['B3'].alignment = center_alignment
    ws['B3'].border = medium_border
    for col in range(3, 5):
        ws.cell(row=3, column=col).border = medium_border

    ws.merge_cells('B4:D4')
    fecha = datetime.now().strftime("%d/%m/%Y")
    ws['B4'] = f"Fecha: {fecha}"
    ws['B4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws['B4'].border = medium_border
    for col in range(3, 5):
        ws.cell(row=4, column=col).border = medium_border

    headers = ['ID', 'Presentación', 'Estado']
    for col_num, header in enumerate(headers, 2):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.fill = green_fill
        cell.font = white_font
        cell.alignment = center_alignment
        cell.border = medium_border

    presentaciones = Presentacion.objects.all()
    for row_num, presentacion in enumerate(presentaciones, 6):  
        ws.cell(row=row_num, column=2, value=presentacion.id)
        ws.cell(row=row_num, column=3, value=f"{presentacion.presentacion} {presentacion.unidad_medida}")
        ws.cell(row=row_num, column=4, value='Activo' if presentacion.estado else 'Inactivo')
        
        for col_num in range(2, 5):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = center_alignment
            cell.border = medium_border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte de presentaciones.xlsx'
    wb.save(response)
    return response

################################################## Productos ##################################################
@login_required
@never_cache
def export_productos_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de productos"

    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    green_fill = PatternFill(start_color="04644B", end_color="04644B", fill_type="solid")
    white_font = Font(color="FFFFFF")
    medium_border = Border(left=Side(style='medium'), 
                         right=Side(style='medium'), 
                         top=Side(style='medium'), 
                         bottom=Side(style='medium'))

    column_width = 20  
    for col in range(2, 10): 
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = column_width

    ws.row_dimensions[2].height = 38 
    ws.row_dimensions[3].height = 23  
    ws.row_dimensions[5].height = 20

    img = Image('app/views/reportes/logo_asuan.png') 
    img.width = 140  
    img.height = 50  
    ws.add_image(img, 'E2')
    ws.merge_cells('B2:I2')  
    ws['B2'].alignment = center_alignment
    ws['B2'].border = medium_border

    ws.merge_cells('B2:I2')
    ws['B2'].alignment = center_alignment
    ws.merge_cells('B3:I3')
    ws['B3'] = "Reporte de Productos"
    ws['B3'].font = Font(size=14, bold=True)
    ws['B3'].alignment = center_alignment
    ws['B3'].border = medium_border
    for col in range(3, 10):
        ws.cell(row=3, column=col).border = medium_border

    ws.merge_cells('B4:I4')
    fecha = datetime.now().strftime("%d/%m/%Y")
    ws['B4'] = f"Fecha: {fecha}"
    ws['B4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws['B4'].border = medium_border
    for col in range(3, 10):
        ws.cell(row=4, column=col).border = medium_border

    headers = ['ID', 'Producto', 'Cantidad', 'Valor', 'Estado', 'Categoría', 'Marca', 'Presentación']
    for col_num, header in enumerate(headers, 2):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.fill = green_fill
        cell.font = white_font
        cell.alignment = center_alignment
        cell.border = medium_border

    productos = Producto.objects.all()
    for row_num, producto in enumerate(productos, 6):  
        ws.cell(row=row_num, column=2, value=producto.id)
        ws.cell(row=row_num, column=3, value=producto.producto)
        ws.cell(row=row_num, column=4, value=producto.cantidad)
        ws.cell(row=row_num, column=5, value=producto.valor)
        ws.cell(row=row_num, column=6, value='Activo' if producto.estado else 'Inactivo')
        ws.cell(row=row_num, column=7, value=producto.id_categoria.categoria)
        ws.cell(row=row_num, column=8, value=producto.id_marca.marca) 
        ws.cell(row=row_num, column=9, value=f"{producto.id_presentacion.presentacion} {producto.id_presentacion.unidad_medida}") 
        
        for col_num in range(2, 10): 
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = center_alignment
            cell.border = medium_border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte de productos.xlsx'
    wb.save(response)
    return response

################################################## Platos ##################################################
@login_required
@never_cache
def export_platos_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de platos"

    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    green_fill = PatternFill(start_color="04644B", end_color="04644B", fill_type="solid")
    white_font = Font(color="FFFFFF")
    medium_border = Border(left=Side(style='medium'), 
                           right=Side(style='medium'), 
                           top=Side(style='medium'), 
                           bottom=Side(style='medium'))

    column_width = 20  
    for col in range(2, 7): 
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = column_width

    ws.row_dimensions[2].height = 38 
    ws.row_dimensions[3].height = 23  
    ws.row_dimensions[4].height = 20

    img = Image('app/views/reportes/logo_asuan.png') 
    img.width = 140  
    img.height = 50  
    ws.add_image(img, 'D2')

    ws.merge_cells('B2:F2')
    ws['B2'] = ""
    ws['B2'].alignment = center_alignment
    ws['B2'].border = medium_border
    ws['C2'].border = medium_border
    ws['D2'].border = medium_border
    ws['E2'].border = medium_border
    ws['F2'].border = medium_border

    ws.merge_cells('B3:F3')
    ws['B3'] = "Reporte de Platos"
    ws['B3'].font = Font(size=14, bold=True)
    ws['B3'].alignment = center_alignment
    ws['B3'].border = medium_border
    ws['C3'].border = medium_border
    ws['D3'].border = medium_border
    ws['E3'].border = medium_border
    ws['F3'].border = medium_border

    ws.merge_cells('B4:F4')
    fecha = datetime.now().strftime("%d/%m/%Y")
    ws['B4'] = f"Fecha: {fecha}"
    ws['B4'].alignment = center_alignment
    ws['B4'].border = medium_border
    for col in range(2, 7):
        ws.cell(row=4, column=col).border = medium_border

    headers = ['ID', 'Plato', 'Descripción', 'Valor', 'Estado']
    for col_num, header in enumerate(headers, 2):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.fill = green_fill
        cell.font = white_font
        cell.alignment = center_alignment
        cell.border = medium_border

    platos = Plato.objects.all()
    for row_num, plato in enumerate(platos, 6):  
        ws.cell(row=row_num, column=2, value=plato.id)
        ws.cell(row=row_num, column=3, value=plato.plato)
        ws.cell(row=row_num, column=4, value=plato.descripcion)
        ws.cell(row=row_num, column=5, value=plato.valor)
        ws.cell(row=row_num, column=6, value='Activo' if plato.estado else 'Inactivo')

        for col_num in range(2, 7):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = center_alignment
            cell.border = medium_border

    for row in ws.iter_rows(min_row=5, min_col=2, max_col=6, max_row=ws.max_row):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = medium_border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte de platos.xlsx'
    wb.save(response)
    return response

################################################## Meseros ##################################################
@login_required
@never_cache
def export_meseros_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de meseros"

    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    green_fill = PatternFill(start_color="04644B", end_color="04644B", fill_type="solid")
    white_font = Font(color="FFFFFF")
    medium_border = Border(left=Side(style='medium'), 
                           right=Side(style='medium'), 
                           top=Side(style='medium'), 
                           bottom=Side(style='medium'))

    column_width = 20  
    for col in range(2, 9): 
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = column_width

    ws.row_dimensions[2].height = 38 
    ws.row_dimensions[3].height = 23  
    ws.row_dimensions[5].height = 20

    img = Image('app/views/reportes/logo_asuan.png') 
    img.width = 140  
    img.height = 50  
    ws.add_image(img, 'E2')
    ws.merge_cells('B2:H2')
    ws['B2'].alignment = center_alignment
    ws['B2'].border = medium_border
    for col in range(3, 9):
        ws.cell(row=2, column=col).border = medium_border

    ws.merge_cells('B3:H3')
    ws['B3'] = "Reporte de Meseros"
    ws['B3'].font = Font(size=14, bold=True)
    ws['B3'].alignment = center_alignment
    ws['B3'].border = medium_border
    for col in range(3, 9):
        ws.cell(row=3, column=col).border = medium_border

    ws.merge_cells('B4:H4')
    fecha = datetime.now().strftime("%d/%m/%Y")
    ws['B4'] = f"Fecha: {fecha}"
    ws['B4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws['B4'].border = medium_border
    for col in range(3, 9):
        ws.cell(row=4, column=col).border = medium_border

    headers = ['ID', 'Nombre', 'Tipo de documento', '# de documento', 'Email', 'Prefijo', 'Teléfono']
    for col_num, header in enumerate(headers, 2):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.fill = green_fill
        cell.font = white_font
        cell.alignment = center_alignment
        cell.border = medium_border

    meseros = Mesero.objects.all()
    for row_num, mesero in enumerate(meseros, 6):  
        ws.cell(row=row_num, column=2, value=mesero.id)
        ws.cell(row=row_num, column=3, value=mesero.nombre)
        ws.cell(row=row_num, column=4, value=mesero.tipo_documento)
        ws.cell(row=row_num, column=5, value=mesero.numero_documento)
        ws.cell(row=row_num, column=6, value=mesero.email)
        ws.cell(row=row_num, column=7, value=mesero.pais_telefono) 
        ws.cell(row=row_num, column=8, value=mesero.telefono)

        for col_num in range(2, 9):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = center_alignment
            cell.border = medium_border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte de meseros.xlsx'
    wb.save(response)
    return response

################################################## Clientes ##################################################
@login_required
@never_cache
def export_clientes_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de clientes"

    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    green_fill = PatternFill(start_color="04644B", end_color="04644B", fill_type="solid")
    white_font = Font(color="FFFFFF")
    medium_border = Border(left=Side(style='medium'), 
                           right=Side(style='medium'), 
                           top=Side(style='medium'), 
                           bottom=Side(style='medium'))

    column_width = 20  
    for col in range(2, 9): 
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = column_width

    ws.row_dimensions[2].height = 38 
    ws.row_dimensions[3].height = 23  
    ws.row_dimensions[5].height = 20

    img = Image('app/views/reportes/logo_asuan.png') 
    img.width = 140  
    img.height = 50  
    ws.add_image(img, 'E2')
    ws.merge_cells('B2:H2')
    ws['B2'].alignment = center_alignment
    ws['B2'].border = medium_border
    for col in range(3, 9):
        ws.cell(row=2, column=col).border = medium_border

    ws.merge_cells('B3:H3')
    ws['B3'] = "Reporte de Clientes"
    ws['B3'].font = Font(size=14, bold=True)
    ws['B3'].alignment = center_alignment
    ws['B3'].border = medium_border
    for col in range(3, 9):
        ws.cell(row=3, column=col).border = medium_border

    ws.merge_cells('B4:H4')
    fecha = datetime.now().strftime("%d/%m/%Y")
    ws['B4'] = f"Fecha: {fecha}"
    ws['B4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws['B4'].border = medium_border
    for col in range(3, 9):
        ws.cell(row=4, column=col).border = medium_border

    headers = ['ID', 'Nombre', 'Tipo de documento', '# de documento', 'Email', 'Prefijo', 'Teléfono']
    for col_num, header in enumerate(headers, 2):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.fill = green_fill
        cell.font = white_font
        cell.alignment = center_alignment
        cell.border = medium_border

    clientes = Cliente.objects.all()
    for row_num, cliente in enumerate(clientes, 6):  
        ws.cell(row=row_num, column=2, value=cliente.id)
        ws.cell(row=row_num, column=3, value=cliente.nombre)
        ws.cell(row=row_num, column=4, value=cliente.tipo_documento)
        ws.cell(row=row_num, column=5, value=cliente.numero_documento)
        ws.cell(row=row_num, column=6, value=cliente.email)
        ws.cell(row=row_num, column=7, value=cliente.pais_telefono) 
        ws.cell(row=row_num, column=8, value=cliente.telefono)

        for col_num in range(2, 9):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = center_alignment
            cell.border = medium_border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte de clientes.xlsx'
    wb.save(response)
    return response

################################################## Administradores ##################################################
@login_required
@never_cache
def export_administradores_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de administradores"

    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    green_fill = PatternFill(start_color="04644B", end_color="04644B", fill_type="solid")
    white_font = Font(color="FFFFFF")
    medium_border = Border(left=Side(style='medium'), 
                           right=Side(style='medium'), 
                           top=Side(style='medium'), 
                           bottom=Side(style='medium'))

    column_width = 20  
    for col in range(2, 8): 
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = column_width

    ws.row_dimensions[2].height = 38 
    ws.row_dimensions[3].height = 23  
    ws.row_dimensions[5].height = 20

    img = Image('app/views/reportes/logo_asuan.png') 
    img.width = 140  
    img.height = 50  
    ws.add_image(img, 'D2')
    ws.merge_cells('B2:G2')
    ws['B2'].alignment = center_alignment
    ws['B2'].border = medium_border
    for col in range(3, 8):
        ws.cell(row=2, column=col).border = medium_border

    ws.merge_cells('B3:G3')
    ws['B3'] = "Reporte de Administradores"
    ws['B3'].font = Font(size=14, bold=True)
    ws['B3'].alignment = center_alignment
    ws['B3'].border = medium_border
    for col in range(3, 8):
        ws.cell(row=3, column=col).border = medium_border

    ws.merge_cells('B4:G4')
    fecha = datetime.now().strftime("%d/%m/%Y")
    ws['B4'] = f"Fecha: {fecha}"
    ws['B4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws['B4'].border = medium_border
    for col in range(3, 8):
        ws.cell(row=4, column=col).border = medium_border

    headers = ['ID', 'Nombre', 'Tipo de documento', '# de documento', 'Email', 'Teléfono']
    for col_num, header in enumerate(headers, 2):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.fill = green_fill
        cell.font = white_font
        cell.alignment = center_alignment
        cell.border = medium_border

    administradores = Administrador.objects.all()
    for row_num, administrador in enumerate(administradores, 6):  
        ws.cell(row=row_num, column=2, value=administrador.id)
        ws.cell(row=row_num, column=3, value=administrador.nombre)
        ws.cell(row=row_num, column=4, value=administrador.tipo_documento)
        ws.cell(row=row_num, column=5, value=administrador.numero_documento)
        ws.cell(row=row_num, column=6, value=administrador.user.email)
        ws.cell(row=row_num, column=7, value=administrador.telefono)

        for col_num in range(2, 8):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = center_alignment
            cell.border = medium_border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte de administradores.xlsx'
    wb.save(response)
    return response

################################################## Operadores ##################################################
@login_required
@never_cache
def export_operadores_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de operadores"

    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    green_fill = PatternFill(start_color="04644B", end_color="04644B", fill_type="solid")
    white_font = Font(color="FFFFFF")
    medium_border = Border(left=Side(style='medium'), 
                           right=Side(style='medium'), 
                           top=Side(style='medium'), 
                           bottom=Side(style='medium'))

    column_width = 20  
    for col in range(2, 8): 
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = column_width

    ws.row_dimensions[2].height = 38 
    ws.row_dimensions[3].height = 23  
    ws.row_dimensions[5].height = 20

    img = Image('app/views/reportes/logo_asuan.png') 
    img.width = 140  
    img.height = 50  
    ws.add_image(img, 'D2')
    ws.merge_cells('B2:G2')
    ws['B2'].alignment = center_alignment
    ws['B2'].border = medium_border
    for col in range(3, 8):
        ws.cell(row=2, column=col).border = medium_border

    ws.merge_cells('B3:G3')
    ws['B3'] = "Reporte de Operadores"
    ws['B3'].font = Font(size=14, bold=True)
    ws['B3'].alignment = center_alignment
    ws['B3'].border = medium_border
    for col in range(3, 8):
        ws.cell(row=3, column=col).border = medium_border

    ws.merge_cells('B4:G4')
    fecha = datetime.now().strftime("%d/%m/%Y")
    ws['B4'] = f"Fecha: {fecha}"
    ws['B4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws['B4'].border = medium_border
    for col in range(3, 8):
        ws.cell(row=4, column=col).border = medium_border

    headers = ['ID', 'Nombre', 'Tipo de documento', '# de documento', 'Email', 'Teléfono']
    for col_num, header in enumerate(headers, 2):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.fill = green_fill
        cell.font = white_font
        cell.alignment = center_alignment
        cell.border = medium_border

    operadores = Operador.objects.all()
    for row_num, operador in enumerate(operadores, 6):  
        ws.cell(row=row_num, column=2, value=operador.id)
        ws.cell(row=row_num, column=3, value=operador.nombre)
        ws.cell(row=row_num, column=4, value=operador.tipo_documento)
        ws.cell(row=row_num, column=5, value=operador.numero_documento)
        ws.cell(row=row_num, column=6, value=operador.user.email)
        ws.cell(row=row_num, column=7, value=operador.telefono)

        for col_num in range(2, 8):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = center_alignment
            cell.border = medium_border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte de operadores.xlsx'
    wb.save(response)
    return response
